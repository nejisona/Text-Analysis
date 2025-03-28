import requests
from bs4 import BeautifulSoup
import openpyxl
import nltk
from nltk.corpus import stopwords
import os
import re
import csv

nltk.download('stopwords')

case = 'lowwer'  

# Download English stop words from NLTK
nltk.download('stopwords')

stop_words = set(stopwords.words('english'))

# Define the directory containing stop word files 
stop_word_dir = "StopWords"  

def read_stop_words(stop_word_dir="StopWords"):
  """
  This function reads stop words from all .txt files in a directory.
  """
  stop_words = set()
  for filename in os.listdir(stop_word_dir):
    if filename.endswith(".txt"):
      with open(os.path.join(stop_word_dir, filename), 'r') as f:
        stop_words.update(line.strip() for line in f)
  return stop_words

# Read positive and negative words from separate text files 
positive_words = set(line.strip().upper() if case == 'upper' else line.strip().lower() for line in open('MasterDictionary/positive-words.txt', 'r'))
negative_words = set(line.strip().upper() if case == 'upper' else line.strip().lower() for line in open('MasterDictionary/negative-words.txt', 'r'))

# # Replace with your actual file paths and sheet names
excel_file = "Input.xlsx"
sheet_name = "Sheet1"
url_column = 2  # Column containing URLs

# Output 
with open('Output.csv', 'w', newline='') as f:
  writer = csv.writer(f)
  header = ["URL ID","URL","POSITIVE SCORE", "NEGATIVE SCORE","POLARITY SCORE","SUBJECTIVITY SCORE" , "Avg Sentence Length", 
            "Percentage Complex Words", "FOG INDEX","Complex Word Count", "WORD COUNT",
            "Syllables per Word", "Personal Pronoun","Avg Word Length"]
  writer.writerow(header)

# Open Excel file
wb = openpyxl.load_workbook(excel_file)
sheet = wb[sheet_name]



def count_complex_words(text):
    vowels = 'aeiouAEIOU'
    complex_word_count = 0
    words = text.lower().split()  # Convert to lowercase and split into words

    for word in words:
        if word.endswith('e') and word[:-1] not in vowels:
            num_syllables = word.count(vowels) - 1  # Exclude final silent 'e'
        else:
            num_syllables = word.count(vowels)
        if num_syllables > 2:
            complex_word_count += 1

    return complex_word_count



def add_attributes_to_csv(attributes, csv_file="Output.csv"):

  with open(csv_file, 'a+') as f:
    writer = csv.writer(f)

    writer.writerow(attributes)

  print(f"Attributes written to {csv_file}")

def count_syllables(word):

    vowels = 'aeiouyAEIOUY'
    vowel_count = 0
    is_first_vowel = True

    for index, char in enumerate(word):
        if char in vowels:
            if is_first_vowel:
                vowel_count += 1
                is_first_vowel = False
            elif index != 0 and word[index-1] not in vowels:
                vowel_count += 1
        else:
            is_first_vowel = True
    if word.endswith('es') and word not in ['wise', 'base', 'use']:
        vowel_count -= 1
    elif word.endswith('ed') and word not in ['bed', 'fired', 'led']:
        vowel_count -= 1

    return vowel_count

def count_personal_pronouns(text):

  pronouns_regex = r"(?i)\b(i|we|my|ours|us(?!A?\b))\b"
  not_us_regex = r"(?i)\b(us\b)"

  clean_text = re.sub(not_us_regex, "", text)


  pronoun_matches = re.findall(pronouns_regex, clean_text)


  personal_pronoun_count = len(pronoun_matches)

  return personal_pronoun_count

def main():

    for row in range(2, sheet.max_row + 1):
        url = sheet.cell(row, url_column).value
        url_ID = sheet.cell(row, url_column-1).value

        try:
            response = requests.get(url)
            response.raise_for_status()

            soup = BeautifulSoup(response.content, "lxml")
        
            data_divs = soup.find_all('div', class_=['td-post-content tagdiv-type', 'tdb-block-inner td-fix-index'])

            if data_divs:
                title = soup.find("h1", class_="entry-title")
                title_text = title.text.strip() if title else ""

                
                body_text = " ".join([p.text.strip() for div in data_divs for p in div.find_all('p')])

                num_words = len(body_text.split())
                sentences = nltk.tokenize.sent_tokenize(body_text)
                num_sentences = len(sentences)

                
                body_text_words = []  
                syllable_counts_words = {}
                syllable_count = 0 
                complex_word_count = 0  

                for div in data_divs:
                    for p in div.find_all('p'):
                        body_text_words.extend(p.text.strip().lower().split())
                        for word in p.text.strip().lower().split():
                            syllable_counts_words[word] = count_syllables(word)
                            syllable_count += count_syllables(word)
                            if syllable_counts_words[word] > 2:
                                complex_word_count += 1

                filtered_words = [word for word in body_text_words if word not in stop_words]

                
                
                num_words_clean = len(filtered_words)
                positive_count = 0
                negative_count = 0
                total_char_count = 0
                average_word_length = total_char_count / num_words
                personal_pronoun_count = count_personal_pronouns(body_text)  

                for word in body_text_words:
                    total_char_count += len(word)

                for word in filtered_words:
                    if word in positive_words:
                        positive_count += 1
                    elif word in negative_words:
                        negative_count += 1
                    
                   
                polarity_score = (positive_count - negative_count) / ((positive_count + negative_count) + 0.000001)
                subjectivity_score = (positive_count - negative_count) / (num_words_clean + 0.000001)

                if num_words > 0:  
                    avg_sentence_length = num_words / num_sentences
                    average_word_length = total_char_count / num_words
                    syllable_per_word = syllable_count /num_words
                    percentage_complex_words = (complex_word_count*100)/num_words
                    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)
                else:
                    average_word_length = 0
                    avg_sentence_length = 0
                
                attributes = [
                    url_ID,
                    url,
                    positive_count, 
                    negative_count,  
                    polarity_score,
                    subjectivity_score,
                    
                    avg_sentence_length, 
                    percentage_complex_words, 
                    fog_index,
                    complex_word_count, 
                    num_words_clean,
                    syllable_per_word, 
                    personal_pronoun_count,  
                    average_word_length 
                ]
                
                add_attributes_to_csv(attributes)
                print(f"Scraped title: {title_text}")
                
            else:
                print(f"No title or body text found for URL: {url}")

        except requests.exceptions.RequestException as e:
            print(f"Error: {e}")


if __name__ == "__main__":
    main()